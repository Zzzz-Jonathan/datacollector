import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook


class DataCollector:
    def __init__(self, urls):
        self.urls = urls
        self.chart_path = ['//*[@id="__next"]/div/div[2]/div[1]/div[3]/div[2]/div[2]/div[2]/div/div/iframe',
                           '//*[@id="sub-container"]/div/div/div[1]/div/div/div[1]/div/div/div/div[2]/div/div[3]']
        # self.chart_path = '/html/body/div[1]/nav/div/ul[1]/li[8]/a'
        self.current_data = []
        self.name = ''

        options = Options()
        options.add_argument("--disable-blink-features")
        options.add_argument("--disable-blink-features=AutomationControlled")

        self.driver = webdriver.Chrome(options=options)
        self.driver.maximize_window()
        self.wait = WebDriverWait(self.driver, 10)
        self.action = ActionChains(self.driver)

        self.wb = Workbook()
        self.ws = self.wb.active
        self.row = 1
        self.col = 1

    def login(self, url, name, password):
        self.driver.get(url)

        try:
            username = self.wait.until(EC.presence_of_element_located((By.XPATH,
                                                                       '//*[@id="__next"]/div/div/div[1]/div/div['
                                                                       '3]/form/div/div[1]/input')))
            userkey = self.wait.until(EC.presence_of_element_located((By.XPATH,
                                                                      '//*[@id="__next"]/div/div/div[1]/div/div['
                                                                      '3]/form/div/div[2]/input')))
            loginbtn = self.wait.until(EC.presence_of_element_located((By.XPATH,
                                                                       '//*[@id="__next"]/div/div/div[1]/div/div['
                                                                       '3]/form/div/button')))

        finally:
            pass

        username.send_keys(name)
        userkey.send_keys(password)

        time.sleep(2)

        loginbtn.click()

    def switch(self, url=None):
        if url is not None:
            self.driver.get(url)

        time.sleep(5)  # wait website loading

        windows = self.driver.window_handles
        self.driver.switch_to.window(windows[-1])

        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, self.chart_path[0]))

    def get_inf(self):
        self.name = self.driver.title
        print(self.name + ' start')
        try:
            charts = self.wait.until(EC.presence_of_element_located((By.XPATH, self.chart_path[1])))
            # charts = self.driver.find_element(By.XPATH, self.chart_path)
        finally:
            pass

        size = charts.size

        for x in range(0, size['width'], 5):
            self.action.move_to_element_with_offset(charts, x, 200).perform()

            try:
                thead = self.driver.find_element(By.XPATH,
                                                 '//*[@id="highcharts-0"]/div/span/div/table/thead/tr/td/span')
                tbody = self.driver.find_element(By.XPATH, '//*[@id="highcharts-0"]/div/span/div/table/tbody')
            except:
                print("Haven't find yet")
            else:
                time_inf = thead.text
                legends = tbody.find_elements(By.CLASS_NAME, 'legend')
                values = tbody.find_elements(By.CLASS_NAME, 'value')
                print(legends[0].text)

                legend_infs = [legend.find_element(By.CLASS_NAME, 'label').text for legend in legends]
                value_infs = [value.find_element(By.TAG_NAME, 'b').text for value in values]

                if time_inf != self.current_data[len(self.current_data) - 1][0]:
                    collect = [time_inf, legend_infs, legend_infs, value_infs]
                    self.current_data.append(collect)

            finally:
                pass

            time.sleep(0.02)

    def toxlsx(self):
        length = len(self.current_data[0][1]) + 1
        col = self.col
        for c in range(col, col + length):
            self.ws.cell(self.row, c).value = self.name
        self.row += 1
        label = 0
        for c in range(col + 1, col + length):
            self.ws.cell(self.row, c).value = self.current_data[0][1][label]
            label += 1
        self.row += 1

        for data in self.current_data:
            self.ws.cell(self.row, col).value = data[0]
            for value, c in zip(data[2], range(col + 1, col + length)):
                self.ws.cell(self.row, c).value = value
            self.row += 1

        self.row = 1
        self.col = col + length

        self.current_data.clear()
        print(self.name + ' finish')


if __name__ == '__main__':
    # basic_inf
    account = 'dengqiaoyun@studio33.wecom.work'
    password = 'NC19980221dqy!'
    links = [
        'https://www.data.ai/apps/ios/app/989673964/rank-history?app_slug=989673964&market_slug=ios&vtype=day'
        '&countries=CN&device=iphone&view=rank&legends=2222&date=2022-02-01~2022-04-18',
        'https://www.data.ai/apps/ios/app/654897098/rank-history?vtype=day&countries=CN&device=iphone&view=grossing'
        '&legends=2222&date=2022-02-01~2022-04-18&market_slug=ios&app_slug=654897098 '
    ]

    dc = DataCollector(links)
    dc.login(dc.urls[1], account, password)  # login data.ai with given user
    time.sleep(5)  # wait website loading
    for app in dc.urls:
        dc.switch(app)
        dc.get_inf()  # moving the mouse and collect inf
        dc.toxlsx()  # push data into xlsx

    dc.wb.save('rank_data.xlsx')
